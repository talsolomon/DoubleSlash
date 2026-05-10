"""
Generates the Duble Slash design-leader interview guide as an Excel workbook.
Run: python3 build_interview_xlsx.py
Output: design-leader-interview-he.xlsx (same folder)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "design-leader-interview-he.xlsx")

DARK = "1a1a2e"
WHITE = "FFFFFF"
STRIPE_A = "f8f9ff"
STRIPE_B = "FFFFFF"
PAIN_BG = "fff3e0"
PAIN_FG = "b45309"
MUST_BG = "ffebee"
HIGH_BG  = "fff8e1"
MED_BG   = "f1f8e9"


def hfill(hex_):
    return PatternFill("solid", fgColor=hex_)

def header_font():
    return Font(name="Arial", bold=True, color=WHITE, size=11)

def body_font(bold=False, color="000000"):
    return Font(name="Arial", bold=bold, size=10, color=color)

def wrap_align(h="right", v="top"):
    return Alignment(wrap_text=True, horizontal=h, vertical=v)

def thin_border():
    s = Side(style="thin", color="d0d0d0")
    return Border(left=s, right=s, top=s, bottom=s)

def set_headers(ws, headers, col_widths):
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    row = ws[1]
    for cell in row:
        cell.fill = hfill(DARK)
        cell.font = header_font()
        cell.alignment = wrap_align(v="center")
        cell.border = thin_border()


wb = Workbook()

# ── Sheet 1: Interview Script ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "סקריפט ראיון"

headers1 = ["חלק", "זמן", "שאלה / הנחיה", "מה לשמוע", "הערות בזמן שיחה"]
widths1  = [18, 9, 52, 36, 28]
set_headers(ws1, headers1, widths1)

script_rows = [
    (
        "לפני השיחה", "—",
        "בדוק את הסטודיו מהרשימה:\n"
        "• גודל הסטודיו (הרכב צוות גס)\n"
        "• סוג העבודה (מיתוג / מוצר / UX / מעורב)\n"
        "• עבודות ציבוריות או לקוחות ששמת לב אליהם\n\n"
        "אל תציג מראש — תן לכאב לצוף באופן טבעי.",
        "—", ""
    ),
    (
        "חימום", "5 דק'",
        "\"לפני שנצלול — תספר לי קצת על הסטודיו? כמה אנשים בצוות, "
        "על איזה סוג עבודה אתם מתמקדים בדרך כלל?\"\n\n"
        "\"ואיך נראה תפקידך ביום-יום — האם אתה בעיקר בתוך העבודה, "
        "או יותר מנהל את הצוות?\"",
        "גודל סטודיו, האם מובל על ידי הבעלים, שיטת תמחור "
        "(שעות / פרויקט / ריטיינר), פרופיל לקוחות.\n\n"
        "אם הם עדיין בתוך העבודה — הם חשים את הכאב של הכלים ישירות.",
        ""
    ),
    (
        "המציאות עם AI", "10 דק'",
        "\"אני רוצה להבין איך AI נחת אצלכם בפועל — לא גרסת ההייפ, "
        "הגרסה האמיתית. האם אתה והצוות משתמשים בזה באופן קבוע?\"\n\n"
        "\"באילו כלים אתם מגיעים הכי הרבה? Claude, ChatGPT, Cursor, "
        "Figma AI — משהו אחר?\"\n\n"
        "\"תסביר לי על הפעם האחרונה שה-AI היה שימושי באמת בפרויקט "
        "אמיתי. איך זה נראה?\"\n\n"
        "\"ומתי זה נשבר? איך נראית עבודת עיצוב עם AI כשהיא מתפרקת?\"",
        "איפה בתהליך (מחקר / קונספט / טקסט / העברה?), סולו או צוות, "
        "באיזה כלי.\n\n"
        "השאלה האחרונה היא המרכזית — תן להם לדבר. אל תמלא שתיקות.",
        ""
    ),
    (
        "בדיקת כאב:\nאובדן הקשר", "10 דק'\n(בחר 2-3\nמתוך 5)",
        "\"כשאתה מתחיל צ'אט חדש על פרויקט שעבדת עליו — כמה אתה "
        "צריך להסביר מחדש? מה זה עולה לך?\"",
        "כאב של רציפות בין-סשן\nתשובת Fish Model: <fish-handoff>",
        ""
    ),
    (
        "בדיקת כאב:\nסחיפת שלב", "",
        "\"קרה לך שביקשת מ-Claude עזרה במחקר או בריף, והוא חזר עם "
        "ווירפריימים או ספק שלא ביקשת? איך אתה מתמודד עם זה?\"",
        "AI עושה יותר מדי, מהר מדי\nתשובת Fish Model: סוכני שלב",
        ""
    ),
    (
        "בדיקת כאב:\nחוסר התאמת עצימות", "",
        "\"האם יש אי-התאמה בין מה שאתה צריך לבין מה שה-AI מפיק "
        "מבחינת היקף? כמו — היית צריך תשובה מהירה והוא כתב עבודת גמר?\"",
        "AI מייצר יותר מדי\nתשובת Fish Model: עצימות לפי סוג הדג",
        ""
    ),
    (
        "בדיקת כאב:\nנראות צוות", "",
        "\"כשהצוות שלך משתמש ב-AI, האם יש לך נראות למה שקורה שם? "
        "או שזה בעצם קופסה שחורה עד שמישהו מראה לך את הפלט?\"",
        "כאב מנהיגות\nתשובת Fish Model: Digest / נראות",
        ""
    ),
    (
        "בדיקת כאב:\nאמון בפלט", "",
        "\"כשה-AI מפיק משהו — טקסט, ספק, המלצה — איך אתה יודע מה "
        "לסמוך עליו? האם יש לך דרך לדעת מה הוא באמת החליט מול מה "
        "שהוא ניחש?\"",
        "אטימות אמון\nתשובת Fish Model: Trust Receipt",
        ""
    ),
    (
        "הפנייה", "2 דק'",
        "ברגע שאתה שומע כאב אמיתי (דוגמה ספציפית, אנחה):\n\n"
        "\"אוכל לשתף משהו שאנחנו בונים? זה קשור ישירות למה שתיארת.\"\n\n"
        "[הפיץ']:\n"
        "\"אנחנו בונים את Duble Slash. זו מתודולוגיה וכלי קל שנותנים "
        "לכלי ה-AI הקשר והמבנה לעבוד כמו שמעצבים עובדים בפועל — שלבים, "
        "העברות, רמת פירוט נכונה לגודל הנכון של הבעיה. הרעיון המרכזי "
        "הוא שה-AI שכבר יש לך הוא מסוגל מספיק — הוא פשוט אין לו מושג "
        "איפה אתה בתהליך או כמה גדולה הבעיה. אנחנו נותנים לו את זה. "
        "התוצאה: סשנים שממשיכים מאיפה שהפסקת, פלטים בגודל הנכון, "
        "והעברות שמעבירות כוונה בפועל.\"\n\n"
        "עצור. תן להם להגיב.\n\n"
        "\"זה נוחת? איזה חלק מדבר אליך הכי הרבה — או לא?\"",
        "האזן לאיזה חלק הם אוחזים — זה מה שתדגיש בהמשך.",
        ""
    ),
    (
        "כשירות לפיילוט", "8 דק'",
        "אם השיחה חמה:\n"
        "\"אנחנו מריצים פיילוט של 8 שבועות עם קבוצה קטנה של סטודיאות "
        "— לגמרי בחינם, עם ליווי צמוד. בתמורה: משוב שבועי ואפשרות "
        "לציין אתכם ברפרנס ברגע שיוצא לאוויר. זה נשמע כמו משהו שתרצה "
        "להיות חלק ממנו?\"\n\n"
        "אם כן:\n"
        "\"האם הוצאה על כלי AI זו החלטה שלך, או שהיא עוברת דרך מישהו "
        "אחר?\"\n"
        "\"האם יש לך אחד-שניים מעצבים שכבר יותר עמוק בשימוש ב-AI?\"\n"
        "\"מה יגרום ל-8 שבועות להרגיש ששווה — מה יצטרך להיות נכון?\"\n\n"
        "אם מהסס:\n"
        "\"מה ההיסוס — האם זה עיתוי, התאמה, משהו לגבי המוצר?\"",
        "בעלים שמקבל החלטות = קונה נכון.\n"
        "מחייב לפי שעה = ROI מהיר ומוחשי.\n"
        "מעצב AI-forward = בודק טבעי.",
        ""
    ),
    (
        "סגירה", "2 דק'",
        "אם מוכשר:\n"
        "\"אשלח לך מסמך קצר שמסביר את הפיילוט בפרטים. אם זה נראה "
        "נכון, נוכל לעשות שיחת המשך קצרה ולאשר משם.\"\n\n"
        "אם לא הרגע הנכון:\n"
        "\"לגמרי בסדר. אוכל לשמור אותך ברשימה ולחזור בעוד חודש "
        "כשיש לנו יותר להראות?\"",
        "תמיד צא עם פעולה הבאה.\nגם \"לא עכשיו\" צריך תאריך.",
        ""
    ),
]

PAIN_ROWS = {4, 5, 6, 7, 8}  # 1-indexed from data start

for i, row in enumerate(script_rows, start=2):
    ws1.append(list(row))
    is_pain = (i - 1) in PAIN_ROWS
    bg = PAIN_BG if is_pain else (STRIPE_A if i % 2 == 0 else STRIPE_B)
    for col in range(1, 6):
        cell = ws1.cell(row=i, column=col)
        cell.fill = hfill(bg)
        cell.alignment = wrap_align()
        cell.border = thin_border()
        if col == 1 and is_pain:
            cell.font = body_font(bold=True, color=PAIN_FG)
        else:
            cell.font = body_font()

# Row heights (approximate)
row_heights = [30, 60, 70, 90, 55, 55, 55, 55, 55, 110, 130, 65]
for i, h in enumerate(row_heights, start=1):
    ws1.row_dimensions[i].height = h

# Freeze header
ws1.freeze_panes = "A2"


# ── Sheet 2: Signal Capture ────────────────────────────────────────────────
ws2 = wb.create_sheet("לכידת סיגנל")

headers2 = [
    "שם הסטודיו", "תאריך", "שם איש קשר",
    "בגרות AI", "כאב מרכזי", "עניין בפיילוט",
    "ציטוט ישיר", "פעולה הבאה"
]
widths2 = [20, 12, 20, 14, 38, 16, 40, 26]
set_headers(ws2, headers2, widths2)

for i in range(2, 12):
    ws2.append([""] * 8)
    bg = STRIPE_A if i % 2 == 0 else STRIPE_B
    for col in range(1, 9):
        cell = ws2.cell(row=i, column=col)
        cell.fill = hfill(bg)
        cell.alignment = wrap_align()
        cell.border = thin_border()
        cell.font = body_font()

ws2.row_dimensions[1].height = 28
for i in range(2, 12):
    ws2.row_dimensions[i].height = 48
ws2.freeze_panes = "A2"


# ── Sheet 3: Qualification Checklist ──────────────────────────────────────
ws3 = wb.create_sheet("כשירות לפיילוט")

headers3 = ["קריטריון כשירות", "עדיפות", "שאלה לשאול", "תשובה"]
widths3 = [34, 12, 42, 30]
set_headers(ws3, headers3, widths3)

qualify_rows = [
    ("הבעלים מקבל החלטות על כלים", "חובה",
     "האם הוצאה על כלי AI זו החלטה שלך?", ""),
    ("הסטודיו מחייב לפי שעות (ROI מהיר)", "גבוהה",
     "האם אתם מחייבים לפי שעה / פרויקט?", ""),
    ("יש מעצב שכבר משתמש ב-AI ביומיום", "גבוהה",
     "מי בצוות הכי עמוק ב-AI כרגע?", ""),
    ("צוות קטן מספיק (2–20 מעצבים)", "בינונית",
     "כמה מעצבים בצוות?", ""),
    ("מוכנות להיות רפרנס ציבורי", "בינונית",
     "נוח לכם שנציין אתכם ברגע שמשחררים?", ""),
    ("מוכנות למשוב שבועי", "בינונית",
     "מה עומס העבודה שלכם ב-8 שבועות הקרובים?", ""),
    ("לא חברת בת של תאגיד גדול", "חובה",
     "(בדוק מראש)", ""),
]

priority_bg = {"חובה": MUST_BG, "גבוהה": HIGH_BG, "בינונית": MED_BG}

for i, row in enumerate(qualify_rows, start=2):
    ws3.append(list(row))
    p = row[1]
    for col in range(1, 5):
        cell = ws3.cell(row=i, column=col)
        cell.alignment = wrap_align()
        cell.border = thin_border()
        cell.font = body_font(bold=(col == 2))
        if col == 2:
            cell.fill = hfill(priority_bg.get(p, WHITE))
        else:
            cell.fill = hfill(STRIPE_A if i % 2 == 0 else STRIPE_B)

ws3.row_dimensions[1].height = 28
for i in range(2, 2 + len(qualify_rows)):
    ws3.row_dimensions[i].height = 44
ws3.freeze_panes = "A2"

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"✅  {OUT}")
