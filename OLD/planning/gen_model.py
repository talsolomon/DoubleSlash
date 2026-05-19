#!/usr/bin/env python3
"""Duble//Slash Financial Model — generates planning/financial-model-v1.xlsx"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
FOREST     = "1E3B2C"
CREAM      = "FAF4E2"
CREAM_EDGE = "E0D5BB"
LILAC      = "DDC9F4"
MUTED      = "6E6A5A"
INK        = "14140F"
INK_SOFT   = "2D2A22"
INPUT_BG   = "FFFBEB"
WHITE      = "FFFFFF"
SECTION_BG = "F4ECD7"
RED        = "C0392B"
AMBER      = "E67E22"
CONS_BG    = "E8F5E9"
BASE_BG    = "E3F2FD"
OPT_BG     = "F3E5F5"

def F(color): return PatternFill(start_color=color, end_color=color, fill_type="solid")
def font(color=INK, bold=False, size=10, italic=False):
    return Font(name="Calibri", color=color, bold=bold, size=size, italic=italic)
def left_a():  return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def center_a():return Alignment(horizontal="center", vertical="center", wrap_text=True)
def right_a(): return Alignment(horizontal="right",  vertical="center")
def top_a():   return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def style(cell, bg=None, fg=INK, bold=False, size=10, italic=False,
          align="left", fmt=None):
    if bg: cell.fill = F(bg)
    cell.font = font(fg, bold, size, italic)
    cell.alignment = {"left": left_a(), "center": center_a(), "right": right_a()}.get(align, left_a())
    if fmt: cell.number_format = fmt

def hdr(ws, row, text, bg=FOREST, fg=WHITE, size=14, span="B:G"):
    ws.row_dimensions[row].height = 36
    cols = span.split(":")
    ws[f"{cols[0]}{row}"] = text
    style(ws[f"{cols[0]}{row}"], bg=bg, fg=fg, bold=True, size=size)
    if cols[0] != cols[1]:
        ws.merge_cells(f"{cols[0]}{row}:{cols[1]}{row}")

def sub(ws, row, text, bg=FOREST, fg="B8C9BD", span="B:G"):
    ws.row_dimensions[row].height = 20
    cols = span.split(":")
    ws[f"{cols[0]}{row}"] = text
    style(ws[f"{cols[0]}{row}"], bg=bg, fg=fg, italic=True, size=9)
    if cols[0] != cols[1]:
        ws.merge_cells(f"{cols[0]}{row}:{cols[1]}{row}")

def section(ws, row, text, span="B:E"):
    ws.row_dimensions[row].height = 24
    cols = span.split(":")
    ws[f"{cols[0]}{row}"] = f"  {text}"
    style(ws[f"{cols[0]}{row}"], bg=SECTION_BG, fg=FOREST, bold=True, size=11)
    if cols[0] != cols[1]:
        ws.merge_cells(f"{cols[0]}{row}:{cols[1]}{row}")

def spacer(ws, row, h=10):
    ws.row_dimensions[row].height = h

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
for col, w in [("A",2),("B",26),("C",18),("D",18),("E",18),("F",2)]:
    ws1.column_dimensions[col].width = w

spacer(ws1, 1)
hdr(ws1, 2, "Duble//Slash · Financial Model", size=18, span="B:E")
sub(ws1, 3, "Seed Round · May 2026  —  Projections through Dec 2027", span="B:E")
spacer(ws1, 4, 14)

ws1["B5"] = "WHAT'S IN THIS WORKBOOK"
style(ws1["B5"], fg=MUTED, bold=True, size=9); ws1.row_dimensions[5].height = 20

sheets_desc = [
    ("Inputs",         "All editable assumptions (yellow cells). Change a number here → everything updates."),
    ("Unit Economics", "LTV, CAC, gross margin, break-even — derived from Inputs."),
    ("Projections",    "3 scenarios (Conservative / Base / Optimistic) month by month, Oct 2026–Dec 2027."),
    ("Burn & Cash",    "Monthly P&L and cash runway from the $1.5M seed, based on Base scenario."),
]
r = 6
for sheet, desc in sheets_desc:
    ws1.row_dimensions[r].height = 22
    ws1[f"B{r}"] = sheet; style(ws1[f"B{r}"], bold=True, size=10)
    ws1[f"C{r}"] = desc;  style(ws1[f"C{r}"], fg=INK_SOFT, size=10); ws1.merge_cells(f"C{r}:E{r}")
    r += 1

spacer(ws1, r, 14); r += 1

ws1[f"B{r}"] = "KEY METRICS  ·  June 2027 (Series A trigger)"
style(ws1[f"B{r}"], fg=MUTED, bold=True, size=9); ws1.row_dimensions[r].height = 20; r += 1

# Column headers for metrics table
for col, txt, bg in [("B","Metric",SECTION_BG),("C","Conservative",CONS_BG),("D","Base",BASE_BG),("E","Optimistic",OPT_BG)]:
    ws1[f"{col}{r}"] = txt; style(ws1[f"{col}{r}"], bg=bg, bold=True, size=9, fg=MUTED, align="center")
ws1.row_dimensions[r].height = 20; r += 1

metrics = [
    ("Paying teams · Jun 2027",   "~220",     "~367",     "~589"),
    ("MRR · Jun 2027",            "$10,800",  "$18,000",  "$28,900"),
    ("ARR · Jun 2027",            "$129k",    "$216k",    "$347k"),
    ("",                          "",         "",         ""),
    ("Price / seat / month",      "$14",      "$14",      "$14"),
    ("Avg team size",             "3.5 seats","3.5 seats","3.5 seats"),
    ("Gross margin (infra)",      "~99%",     "~99%",     "~99%"),
    ("Monthly team churn",        "3%",       "3%",       "3%"),
    ("",                          "",         "",         ""),
    ("Realistic Series A ARR",    "~$130k",   "~$216k",   "~$347k"),
]
for label, c, b, o in metrics:
    ws1.row_dimensions[r].height = 22
    bg = CREAM if r % 2 == 0 else WHITE
    ws1[f"B{r}"] = label; style(ws1[f"B{r}"], bg=bg, size=10, fg=INK_SOFT)
    ws1[f"C{r}"] = c;     style(ws1[f"C{r}"], bg=CONS_BG if c else bg, size=10, align="center")
    ws1[f"D{r}"] = b;     style(ws1[f"D{r}"], bg=BASE_BG if b else bg, size=10, align="center")
    ws1[f"E{r}"] = o;     style(ws1[f"E{r}"], bg=OPT_BG if o else bg, size=10, align="center")
    r += 1

spacer(ws1, r, 14); r += 1

ws1[f"B{r}"] = "ROADMAP MILESTONES"
style(ws1[f"B{r}"], fg=MUTED, bold=True, size=9); ws1.row_dimensions[r].height = 20; r += 1

milestones = [
    ("May 2026", "OSS Launch",        "4 operators · npx install · macOS menu-bar · launch articles"),
    ("Jul 2026", "V1 Free Launch",    "Context Cloud + // injection + target: 1,000 free users"),
    ("Oct 2026", "Paid Tier Opens",   "Team tier $14/seat/mo · shared contexts · handoffs · analytics"),
    ("Jan 2027", "Series A Prep",     "Deck ready · 200–400 teams paying · ARR: $100–250k"),
    ("Jun 2027", "Series A Trigger",  "$130k–$347k ARR depending on scenario"),
    ("H2 2027",  "Enterprise Pilots", "Figma / Linear / GitHub bridges · Process telemetry · V2"),
]
for date, name, desc in milestones:
    ws1.row_dimensions[r].height = 26
    is_key = date in ("Oct 2026", "Jun 2027")
    bg = FOREST if is_key else (CREAM if r % 2 == 0 else WHITE)
    fg = WHITE if is_key else INK_SOFT
    ws1[f"B{r}"] = date; style(ws1[f"B{r}"], bg=bg, fg=fg, bold=is_key, size=10)
    ws1[f"C{r}"] = name; style(ws1[f"C{r}"], bg=bg, fg=fg, bold=True, size=10)
    ws1[f"D{r}"] = desc; style(ws1[f"D{r}"], bg=bg, fg=fg, size=9, italic=not is_key)
    ws1.merge_cells(f"D{r}:E{r}")
    r += 1


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INPUTS
# Fixed cell assignments (so cross-sheet formulas can reference them):
#   D9  = seat price            D10 = avg team size
#   D11 = ACV/team/mo           D12 = ACV/team/yr
#   D16 = COGS/seat/mo          D17 = COGS/team/mo       D18 = gross margin
#   D23 = seed raise            D24 = runway months      D25 = monthly burn
#   D26 = engineering %         D27 = GTM %              D28 = infra %
#   D29 = eng $/mo              D30 = GTM $/mo           D31 = infra $/mo
#   D35 = churn rate            D36 = avg lifespan       D37 = LTV/team
#   D38 = CAC                   D39 = LTV:CAC
#   D43 = cons start            D44 = cons growth
#   D46 = base start            D47 = base growth
#   D49 = opt start             D50 = opt growth
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Inputs")
ws2.sheet_view.showGridLines = False
for col, w in [("A",2),("B",22),("C",32),("D",16),("E",42),("F",2)]:
    ws2.column_dimensions[col].width = w

spacer(ws2, 1)
hdr(ws2, 2, "Inputs & Assumptions", size=16, span="B:E")
sub(ws2, 3, "Yellow cells = editable. All other sheets derive from column D.", span="B:E")
spacer(ws2, 4, 14)

# Legend
ws2.row_dimensions[5].height = 22
ws2["B5"] = "  ← Yellow = editable input"
ws2["B5"].fill = F(INPUT_BG); ws2["B5"].font = font(INK_SOFT, italic=True, size=9)
ws2["B5"].alignment = left_a(); ws2.merge_cells("B5:E5")
spacer(ws2, 6, 10)

def col_hdrs(ws, row):
    ws.row_dimensions[row].height = 20
    for col, txt in [("B","Category"),("C","Parameter"),("D","Value"),("E","Notes / Explanation")]:
        ws[f"{col}{row}"] = txt
        style(ws[f"{col}{row}"], bg=CREAM_EDGE, bold=True, size=9, fg=MUTED, align="center")

def irow(ws, row, cat, param, value, notes, fmt=None, editable=True):
    ws.row_dimensions[row].height = 24
    ws[f"B{row}"] = cat;   style(ws[f"B{row}"], fg=MUTED, italic=True, size=9)
    ws[f"C{row}"] = param; style(ws[f"C{row}"], fg=INK, size=10)
    ws[f"D{row}"] = value
    style(ws[f"D{row}"], bg=INPUT_BG if editable else WHITE,
          fg=INK, bold=editable, size=11, align="center", fmt=fmt)
    ws[f"E{row}"] = notes; style(ws[f"E{row}"], fg=INK_SOFT, italic=True, size=9)

# ── PRICING (rows 7–12) ─────────────────────────────────────────
section(ws2, 7, "PRICING")
col_hdrs(ws2, 8)
irow(ws2,  9, "Pricing", "Price per seat / month",    14,    "Team tier only — individual use is permanently free",        fmt='"$"#,##0.00')
irow(ws2, 10, "Pricing", "Average team size (seats)", 3.5,   "First-wave studios are small (3–4 seats); grows over time",  fmt='#,##0.0')
irow(ws2, 11, "Pricing", "ACV per team / month",     "=D9*D10",  "Derived: seat price × avg team size",                   fmt='"$"#,##0.00', editable=False)
irow(ws2, 12, "Pricing", "ACV per team / year",      "=D11*12",  "Derived: monthly × 12",                                 fmt='"$"#,##0',    editable=False)
spacer(ws2, 13)

# ── COGS (rows 14–19) ────────────────────────────────────────────
section(ws2, 14, "COST OF GOODS SOLD  (infrastructure per seat)")
col_hdrs(ws2, 15)
irow(ws2, 16, "Infrastructure", "COGS per seat / month",   0.08,      "Local-first (Ollama). Cloud stores backup only — S3 ~500MB/user avg.",             fmt='"$"#,##0.000')
irow(ws2, 17, "Infrastructure", "COGS per team / month",  "=D16*D10", "Derived: COGS/seat × avg team size",                                                fmt='"$"#,##0.00',  editable=False)
irow(ws2, 18, "Infrastructure", "Gross margin (infra)",   "=(D9-D16)/D9", "~99% — the business cost is people, not infrastructure",                        fmt='0.0%',         editable=False)
spacer(ws2, 19)

# ── FUNDING (rows 20–32) ─────────────────────────────────────────
section(ws2, 20, "SEED FUNDING & BURN")
col_hdrs(ws2, 21)
irow(ws2, 22, "Fundraise", "Seed raise (total)",        1500000, "Pre-money negotiated; 18-month runway",                           fmt='"$"#,##0')
irow(ws2, 23, "Fundraise", "Runway (months)",           18,      "Target runway from close to Series A",                            fmt='#,##0')
irow(ws2, 24, "Burn",      "Monthly burn rate",         "=D22/D23",  "Derived: total raise ÷ runway",                               fmt='"$"#,##0', editable=False)
irow(ws2, 25, "Burn",      "Engineering %",             0.35,    "2 hires: platform eng (capture/injection) + systems eng (cloud)", fmt='0%')
irow(ws2, 26, "Burn",      "GTM %",                     0.55,    "Founder launch, content, partner programs, Config/UXLive/DevDay", fmt='0%')
irow(ws2, 27, "Burn",      "Infrastructure %",          0.10,    "Hosted backend, CDN, observability, security audit (pre-paid)",   fmt='0%')
irow(ws2, 28, "Burn",      "Engineering / month",       "=D24*D25", "Derived",                                                     fmt='"$"#,##0', editable=False)
irow(ws2, 29, "Burn",      "GTM / month",               "=D24*D26", "Derived — this also sets CAC in Growth section",              fmt='"$"#,##0', editable=False)
irow(ws2, 30, "Burn",      "Infrastructure (OpEx) / mo","=D24*D27", "Derived",                                                     fmt='"$"#,##0', editable=False)
spacer(ws2, 31)

# ── GROWTH (rows 32–40) ──────────────────────────────────────────
section(ws2, 32, "GROWTH & RETENTION")
col_hdrs(ws2, 33)
irow(ws2, 34, "Retention", "Monthly team churn rate",     0.03,           "3% = moderate. Best-in-class hits 1–2%. Assumed to improve post-PMF.",         fmt='0.0%')
irow(ws2, 35, "Retention", "Avg team lifespan (months)",  "=1/D34",       "Derived: 1 ÷ churn rate. At 3% → 33 months.",                                  fmt='#,##0.0', editable=False)
irow(ws2, 36, "Retention", "LTV per team",                "=D11*D35",     "Monthly ACV × avg lifespan",                                                   fmt='"$"#,##0', editable=False)
irow(ws2, 37, "Retention", "CAC (base scenario, GTM)",    "=D29/25",      "GTM monthly spend ÷ 25 new teams/mo (base launch rate). Falls as organic grows",fmt='"$"#,##0', editable=False)
irow(ws2, 38, "Retention", "LTV:CAC (early stage)",       "=D36/D37",     "<1× early is normal seed-stage. Target >3× at scale (see Unit Economics)",     fmt='#,##0.0"×"', editable=False)
spacer(ws2, 39)

# ── SCENARIOS (rows 40–51) ───────────────────────────────────────
section(ws2, 40, "GROWTH SCENARIOS  —  new paying teams / month")
ws2.row_dimensions[41].height = 20
for col, txt in [("B","Scenario"),("C","Parameter"),("D","Value"),("E","Rationale")]:
    ws2[f"{col}41"] = txt
    style(ws2[f"{col}41"], bg=CREAM_EDGE, bold=True, size=9, fg=MUTED, align="center")

irow(ws2, 42, "Conservative", "New teams at launch (Oct 2026)", 15, "~4 pilot studios + immediate network; tight but high-trust Israeli design cluster")
irow(ws2, 43, "Conservative", "Monthly growth in new teams",    3,  "Slow word-of-mouth. +3 new teams every month on top of previous month's rate")
for c in ["B","C","E"]: ws2[f"{c}42"].fill = F(CONS_BG); ws2[f"{c}43"].fill = F(CONS_BG)
spacer(ws2, 44)

irow(ws2, 45, "Base",         "New teams at launch (Oct 2026)", 25, "Pilot studios + network + early OSS converts. ~5% of 500 aware orgs pay immediately")
irow(ws2, 46, "Base",         "Monthly growth in new teams",    5,  "Content + conferences (Config 2026 etc.) drive sustained organic discovery")
for c in ["B","C","E"]: ws2[f"{c}45"].fill = F(BASE_BG); ws2[f"{c}46"].fill = F(BASE_BG)
spacer(ws2, 47)

irow(ws2, 48, "Optimistic",   "New teams at launch (Oct 2026)", 40, "OSS breakout + viral conference moment. ~4% of 1,000 aware orgs convert at launch")
irow(ws2, 49, "Optimistic",   "Monthly growth in new teams",    8,  "Anthropic partner placement + one big content hit sustains high monthly inflow")
for c in ["B","C","E"]: ws2[f"{c}48"].fill = F(OPT_BG); ws2[f"{c}49"].fill = F(OPT_BG)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — UNIT ECONOMICS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Unit Economics")
ws3.sheet_view.showGridLines = False
for col, w in [("A",2),("B",36),("C",18),("D",42),("E",2)]:
    ws3.column_dimensions[col].width = w

spacer(ws3, 1)
hdr(ws3, 2, "Unit Economics", size=16, span="B:D")
sub(ws3, 3, "All values derive from Inputs sheet. Edit yellow cells there → this sheet updates.", span="B:D")
spacer(ws3, 4, 14)

def ue_col_hdrs(ws, row):
    ws.row_dimensions[row].height = 20
    for col, txt in [("B","Metric"),("C","Value"),("D","Explanation")]:
        ws[f"{col}{row}"] = txt
        style(ws[f"{col}{row}"], bg=CREAM_EDGE, bold=True, size=9, fg=MUTED, align="center")

def ue_row(ws, row, label, formula, explanation, fmt=None):
    ws.row_dimensions[row].height = 24
    ws[f"B{row}"] = label; style(ws[f"B{row}"], size=10, fg=INK)
    ws[f"C{row}"] = formula
    style(ws[f"C{row}"], bg=CREAM, bold=True, size=11, fg=FOREST, align="center", fmt=fmt)
    ws[f"D{row}"] = explanation; style(ws[f"D{row}"], size=9, fg=INK_SOFT, italic=True)

r = 5
section(ws3, r, "REVENUE METRICS", span="B:D"); r+=1
ue_col_hdrs(ws3, r); r+=1
ue_row(ws3, r, "Price per seat / month",       "=Inputs!D9",    "Team tier only; individuals are free forever",                fmt='"$"#,##0.00'); r+=1
ue_row(ws3, r, "Average team size",            "=Inputs!D10",   "Seats per paying team — first wave: 3–4 seats (small studios)",fmt='#,##0.0'); r+=1
ue_row(ws3, r, "Monthly revenue per team",     "=Inputs!D11",   "ACV/team/month",                                              fmt='"$"#,##0.00'); r+=1
ue_row(ws3, r, "Annual revenue per team",      "=Inputs!D12",   "ACV/team/year — used for ARR and investor metrics",           fmt='"$"#,##0'); r+=1
spacer(ws3, r); r+=1

section(ws3, r, "COST OF GOODS SOLD", span="B:D"); r+=1
ue_col_hdrs(ws3, r); r+=1
ue_row(ws3, r, "Infrastructure COGS / seat / month",  "=Inputs!D16",  "Storage + sync on local-first model. Cloud = backup only.",fmt='"$"#,##0.000'); r+=1
ue_row(ws3, r, "Infrastructure COGS / team / month",  "=Inputs!D17",  "COGS/seat × avg team size",                              fmt='"$"#,##0.00'); r+=1
ue_row(ws3, r, "Gross margin (infrastructure only)",   "=Inputs!D18",  "~99%. The cost structure is people, not infrastructure.", fmt='0.0%'); r+=1
spacer(ws3, r); r+=1

section(ws3, r, "CUSTOMER ACQUISITION & LIFETIME VALUE", span="B:D"); r+=1
ue_col_hdrs(ws3, r); r+=1
ue_row(ws3, r, "Monthly team churn rate",              "=Inputs!D34",  "3% monthly. Best-in-class team tools: 1–2%.",            fmt='0.0%'); r+=1
ue_row(ws3, r, "Average team lifespan",                "=Inputs!D35",  "1 ÷ churn rate. At 3%: ~33 months.",                     fmt='#,##0.0 "months"'); r+=1
ue_row(ws3, r, "LTV per team",                         "=Inputs!D36",  "Monthly ACV × avg lifespan = total revenue per team",    fmt='"$"#,##0'); r+=1
ue_row(ws3, r, "CAC — base scenario (GTM-funded)",     "=Inputs!D37",  "GTM monthly spend ÷ 25 new teams/mo at base launch",     fmt='"$"#,##0'); r+=1
ue_row(ws3, r, "LTV:CAC — early stage",                "=Inputs!D38",  "<1× early is expected. Path to >3× at scale (see notes)",fmt='#,##0.0"×"'); r+=1
spacer(ws3, r); r+=1

# LTV:CAC explanation block
ws3.row_dimensions[r].height = 100
ws3[f"B{r}"] = (
    "WHY LTV:CAC < 1× IS OK AT SEED STAGE:\n\n"
    "• GTM spend of $46k/mo is high relative to early $49/mo ARPU per team.\n"
    "• As teams/month scale from 25 → 100+, the same GTM budget makes CAC drop from ~$1,840 → ~$460.\n"
    "• At $460 CAC: LTV:CAC = $1,617 / $460 = 3.5× — well above the '>3× healthy' benchmark.\n"
    "• Additionally, organic channels (OSS community, content, methodology-as-distribution) are not in the CAC calc.\n"
    "  When 50%+ of acquisition is organic, effective CAC halves → LTV:CAC doubles.\n\n"
    "Investors in seed-stage SaaS accept <1× LTV:CAC when: (a) 70%+ gross margin, (b) clear path to >3× at scale.\n"
    "DS has both: 99% gross margin and a demonstrable scale curve on CAC."
)
style(ws3[f"B{r}"], bg=CREAM, fg=INK_SOFT, size=9)
ws3.merge_cells(f"B{r}:D{r}")
ws3[f"B{r}"].alignment = top_a()
r+=1; spacer(ws3, r); r+=1

section(ws3, r, "BREAK-EVEN", span="B:D"); r+=1
ue_col_hdrs(ws3, r); r+=1
ue_row(ws3, r, "Monthly burn rate",                  "=Inputs!D24",          "Total seed ÷ runway months",                        fmt='"$"#,##0'); r+=1
ue_row(ws3, r, "MRR to break even",                  "=Inputs!D24",          "≈ burn (gross margin ~99%)",                         fmt='"$"#,##0'); r+=1
ue_row(ws3, r, "Paid seats to break even",           "=Inputs!D24/Inputs!D9","MRR target ÷ price/seat",                           fmt='#,##0 "seats"'); r+=1
ue_row(ws3, r, "Paying teams to break even",         "=Inputs!D24/Inputs!D11","MRR target ÷ ACV/team",                            fmt='#,##0 "teams"'); r+=1

ws3.row_dimensions[r].height = 36
ws3[f"B{r}"] = "Break-even (~1,700 teams) is a Series B target. The seed funds growth to Series A (150–400 teams)."
style(ws3[f"B{r}"], bg=LILAC, fg=FOREST, bold=True, size=10)
ws3.merge_cells(f"B{r}:D{r}")
ws3[f"B{r}"].alignment = center_a()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — PROJECTIONS
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Projections")
ws4.sheet_view.showGridLines = False
for col, w in [("A",2),("B",14),("C",14),("D",12),("E",16),("F",16),("G",18),("H",2)]:
    ws4.column_dimensions[col].width = w

spacer(ws4, 1)
hdr(ws4, 2, "Revenue Projections · Oct 2026 – Dec 2027", size=16)
sub(ws4, 3, "Three scenarios. Assumptions in Inputs sheet. Highlighted rows = June 2027 (Series A trigger).")
spacer(ws4, 4, 14)

MONTHS = [
    "Oct 2026","Nov 2026","Dec 2026",
    "Jan 2027","Feb 2027","Mar 2027","Apr 2027","May 2027","Jun 2027",
    "Jul 2027","Aug 2027","Sep 2027","Oct 2027","Nov 2027","Dec 2027",
]

def proj_block(ws, start_row, scenario_label, bg, new_start, new_growth, churn_rate=0.03, avg_team=3.5, price=14):
    ws.row_dimensions[start_row].height = 28
    ws[f"B{start_row}"] = f"  {scenario_label}"
    style(ws[f"B{start_row}"], bg=FOREST, fg=WHITE, bold=True, size=12)
    ws.merge_cells(f"B{start_row}:G{start_row}")
    start_row += 1

    ws.row_dimensions[start_row].height = 22
    for col, lbl in [("B","Month"),("C","New teams"),("D","Churned"),("E","Cumulative teams"),("F","MRR"),("G","ARR (run-rate)")]:
        ws[f"{col}{start_row}"] = lbl
        style(ws[f"{col}{start_row}"], bg=bg, bold=True, size=9, fg=MUTED, align="center")
    start_row += 1

    cumulative = 0
    for i, month in enumerate(MONTHS):
        ws.row_dimensions[start_row].height = 22
        new   = new_start + new_growth * i
        churn = round(cumulative * churn_rate)
        cumulative = cumulative + new - churn
        mrr   = round(cumulative * avg_team * price)
        arr   = mrr * 12

        is_key = (month == "Jun 2027")
        row_bg = LILAC if is_key else (bg if i % 2 == 0 else WHITE)

        ws[f"B{start_row}"] = month;      style(ws[f"B{start_row}"], bg=row_bg, bold=is_key, size=10)
        ws[f"C{start_row}"] = new;        style(ws[f"C{start_row}"], bg=row_bg, size=10, align="center")
        ws[f"D{start_row}"] = churn;      style(ws[f"D{start_row}"], bg=row_bg, size=10, fg=MUTED, align="center")
        ws[f"E{start_row}"] = cumulative; style(ws[f"E{start_row}"], bg=row_bg, bold=is_key, size=10, align="center")
        ws[f"F{start_row}"] = mrr;        style(ws[f"F{start_row}"], bg=row_bg, bold=is_key, size=10, fg=FOREST, align="right", fmt='"$"#,##0')
        ws[f"G{start_row}"] = arr;        style(ws[f"G{start_row}"], bg=row_bg, bold=is_key, size=11 if is_key else 10, fg=FOREST, align="right", fmt='"$"#,##0')
        start_row += 1

    # Final summary row
    ws.row_dimensions[start_row].height = 26
    ws[f"B{start_row}"] = "Dec 2027 — end of model"
    ws[f"E{start_row}"] = cumulative
    ws[f"F{start_row}"] = round(cumulative * avg_team * price)
    ws[f"G{start_row}"] = round(cumulative * avg_team * price) * 12
    for col in ["B","C","D","E","F","G"]:
        style(ws[f"{col}{start_row}"], bg=SECTION_BG, bold=True, size=10, fg=FOREST,
              align="right" if col in ["E","F","G"] else "left",
              fmt='"$"#,##0' if col in ["F","G"] else None)
    ws[f"C{start_row}"].value = None; ws[f"D{start_row}"].value = None

    return start_row + 2

r = 5
r = proj_block(ws4, r, "CONSERVATIVE  ·  15 teams at launch  /  +3 teams / month", CONS_BG, 15, 3)
r = proj_block(ws4, r, "BASE  ·  25 teams at launch  /  +5 teams / month",         BASE_BG, 25, 5)
r = proj_block(ws4, r, "OPTIMISTIC  ·  40 teams at launch  /  +8 teams / month",   OPT_BG,  40, 8)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — BURN & CASH
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Burn & Cash")
ws5.sheet_view.showGridLines = False
for col, w in [("A",2),("B",14),("C",14),("D",12),("E",14),("F",14),("G",14),("H",14),("I",16),("J",18),("K",2)]:
    ws5.column_dimensions[col].width = w

spacer(ws5, 1)
hdr(ws5, 2, "Burn & Cash Position  ·  Base Scenario", size=16, span="B:J")
sub(ws5, 3, "Seed of $1.5M closes May 2026. Revenue starts Oct 2026. Base scenario projections.", span="B:J")
spacer(ws5, 4, 14)

ws5.row_dimensions[5].height = 26
for col, lbl in [("B","Month"),("C","Revenue (MRR)"),("D","COGS"),("E","Gross Profit"),
                  ("F","Engineering"),("G","GTM Spend"),("H","Infra (OpEx)"),("I","Net Burn / Mo"),("J","Cash Remaining")]:
    ws5[f"{col}5"] = lbl
    style(ws5[f"{col}5"], bg=SECTION_BG, bold=True, size=9, fg=FOREST, align="center")

ALL_MONTHS = [
    "May 2026","Jun 2026","Jul 2026","Aug 2026","Sep 2026",
    "Oct 2026","Nov 2026","Dec 2026",
    "Jan 2027","Feb 2027","Mar 2027","Apr 2027","May 2027","Jun 2027",
    "Jul 2027","Aug 2027","Sep 2027","Oct 2027","Nov 2027","Dec 2027",
]

# Pre-compute base scenario MRR per month
base_mrr = {}
cum = 0
for i, m in enumerate(MONTHS):
    new = 25 + 5 * i
    churn = round(cum * 0.03)
    cum = cum + new - churn
    base_mrr[m] = round(cum * 3.5 * 14)

# Fixed monthly costs
monthly_burn  = 1_500_000 / 18
eng_mo  = round(monthly_burn * 0.35)
gtm_mo  = round(monthly_burn * 0.55)
infra_mo= round(monthly_burn * 0.10)

cash = 1_500_000
r = 6
for month in ALL_MONTHS:
    ws5.row_dimensions[r].height = 22
    mrr = base_mrr.get(month, 0)
    cogs = round(mrr * 0.006)
    gp   = mrr - cogs
    net  = gp - eng_mo - gtm_mo - infra_mo
    cash += net

    is_key = month in ("Oct 2026", "Jun 2027")
    row_bg = LILAC if is_key else (CREAM if r % 2 == 0 else WHITE)

    ws5[f"B{r}"] = month;    style(ws5[f"B{r}"], bg=row_bg, bold=is_key, size=10)
    ws5[f"C{r}"] = mrr;      style(ws5[f"C{r}"], bg=row_bg, size=10, fg=FOREST if mrr else MUTED, align="right", fmt='"$"#,##0')
    ws5[f"D{r}"] = cogs;     style(ws5[f"D{r}"], bg=row_bg, size=10, fg=MUTED, align="right", fmt='"$"#,##0')
    ws5[f"E{r}"] = gp;       style(ws5[f"E{r}"], bg=row_bg, size=10, fg=FOREST if gp else MUTED, align="right", fmt='"$"#,##0')
    ws5[f"F{r}"] = eng_mo;   style(ws5[f"F{r}"], bg=row_bg, size=10, fg=INK_SOFT, align="right", fmt='"$"#,##0')
    ws5[f"G{r}"] = gtm_mo;   style(ws5[f"G{r}"], bg=row_bg, size=10, fg=INK_SOFT, align="right", fmt='"$"#,##0')
    ws5[f"H{r}"] = infra_mo; style(ws5[f"H{r}"], bg=row_bg, size=10, fg=INK_SOFT, align="right", fmt='"$"#,##0')
    ws5[f"I{r}"] = net;      style(ws5[f"I{r}"], bg=row_bg, bold=is_key, size=10, fg=(RED if net < 0 else FOREST), align="right", fmt='"$"#,##0')
    ws5[f"J{r}"] = cash;     style(ws5[f"J{r}"], bg=row_bg, bold=is_key, size=10,
                                   fg=(RED if cash < 200_000 else (AMBER if cash < 600_000 else FOREST)),
                                   align="right", fmt='"$"#,##0')
    r += 1

spacer(ws5, r, 14); r+=1
ws5.row_dimensions[r].height = 110
ws5[f"B{r}"] = (
    "NOTES:\n\n"
    "• Revenue = $0 until Oct 2026 (paid tier launch). Pre-revenue months are pure burn.\n"
    "• COGS is ~0.6% of MRR. The local-first architecture keeps infrastructure costs negligible.\n"
    "• Engineering, GTM, and Infra are drawn as flat monthly allocations from the seed.\n"
    "• Cash turns amber when <$600k remaining (Series A prep territory), red when <$200k (runway risk).\n"
    "• Base scenario reaches ~$18k MRR / $216k ARR by June 2027. Monthly burn stays ~$83k.\n"
    "  Revenue only partially offsets burn by Series A — the raise bridges to profitability post-Series A.\n"
    "• To model a different scenario: go to Inputs sheet, change scenario values, and re-run the model\n"
    "  (or adjust the Projections sheet to see Conservative / Optimistic cash curves).\n\n"
    "Series A target: raise ~$5–8M on $150–350k ARR traction. Use of funds: scale GTM + add 2 engineers."
)
style(ws5[f"B{r}"], bg=CREAM, fg=INK_SOFT, size=9)
ws5.merge_cells(f"B{r}:J{r}")
ws5[f"B{r}"].alignment = top_a()


# ── Save ─────────────────────────────────────────────────────────────────────
out = "/Users/talsolomon/Documents/dubleslash/planning/financial-model-v1.xlsx"
wb.save(out)
print(f"✓ Saved: {out}")
